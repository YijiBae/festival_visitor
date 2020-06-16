from django.shortcuts import render
from fes_visitor.forms import Form
from fes_visitor.forms import Option, Results
import pandas as pd
from pandas import DataFrame as df

import datetime
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
import requests
from bs4 import BeautifulSoup
from selenium import webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
from openpyxl import load_workbook
from collections import Counter
import os

def calc_day(start_year, start_month, start_day, end_year, end_month, end_day):
    start_date = datetime.datetime(start_year, start_month, start_day)
    end_date = datetime.datetime(end_year, end_month, end_day)
    day_term = (end_date - start_date).days
    return day_term

def calc_else(fes_name):
    #엑셀파일 불러오기
    count = 0
    lists = load_workbook('C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/통합축제데이터.xlsx')
    sheet = lists['Sheet1']
    for i in sheet: 
        count += 1
        #제일 첫 열이 None이라서 넘어가는 것
        if(count > 1):                
            if(i[3].value == fes_name):
                KTX = i[11].value
                local_code = i[28].value
                neutral = i[21].value
                positive = i[22].value
                break
    return KTX, local_code, neutral, positive
def regression_model(test_value):
    festival_df = pd.read_excel('C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/통합축제데이터.xlsx', header = 0, sheet_name = 'Sheet1')
    #있는 것 만으로 감정분석
    sentiment = festival_df.dropna(axis=0)
    x_train = sentiment[['기간', "KTX역",'평균기온 절대값', '최저기온 절대값', '최고기온 절대값','평균운량', 'positive', 'neutral']]
    y_train = sentiment[['합계']]
    
    mlr= LinearRegression()
    mlr.fit(x_train, y_train)
    y_predict = mlr.predict(test_value)
    return y_predict

def bass_model(fes_name, start_year):
    lists = load_workbook('C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/연속누적방문객.xlsx')
    sheet = lists['BASS']
    count = 0
    check = -1
    for i in sheet: 
        count += 1
        #제일 첫 열이 None이라서 넘어가는 것
        if(count > 1):
            if(i[0].value == fes_name):
                #년도 찾기
                if(start_year == 2007):
                    return (i[1].value)
                    check = 1 
                elif(start_year== 2008):
                    return (i[2].value)
                    check = 1 
                elif(start_year == 2009):
                    return (i[3].value)
                    check = 1 
                elif(start_year == 2010):
                    return (i[4].value)
                    check = 1 
                elif(start_year == 2011):
                    return (i[5].value)
                    check = 1 
                elif(start_year== 2012):
                    return (i[6].value)
                    check = 1 
                elif(start_year == 2013):
                    return (i[7].value)
                    check = 1 
                elif(start_year == 2014):
                    return (i[8].value)
                    check = 1 
                elif(start_year == 2015):
                    return (i[9].value)
                    check = 1 
                elif(start_year == 2016):
                    return (i[10].value)
                    check = 1 
                elif(start_year == 2017):
                    return (i[11].value)
                    check = 1 
                elif(start_year == 2018):
                    return (i[12].value)
                    check = 1 
                elif(start_year == 2019):
                    return (i[13].value)
                    check = 1 
                elif(start_year == 2020):
                    return (i[14].value)
                    check = 1 
                elif(start_year == 2021):
                    return (i[15].value)
                    check = 1 
                elif(start_year == 2022):
                    return (i[16].value)
                    check = 1 
                elif(start_year == 2023):
                    return (i[17].value)
                    check = 1 
                elif(start_year == 2024):
                    return (i[18].value)
                    check = 1 
                elif(start_year == 2025):
                    return (i[19].value)
                    check = 1 
                elif(start_year == 2026):
                    return (i[20].value)
                    check = 1 
                elif(start_year == 2027):
                    return (i[21].value)
                    check = 1 
                elif(start_year == 2028):
                    return (i[22].value)
                    check = 1 
                break

def hybrid_model(regression_result, bass_result):
    final_result = (0.6*bass_result + 0.4* regression_result)
    return final_result


def draw_visitor(fes_name):
    lists = load_workbook('C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/연속누적방문객.xlsx')
    sheet = lists['Sheet1']
    count = 0
    year = np.array([2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017])
    check = -1
    for i in sheet: 
        count += 1
        if(count > 1):
            if(i[0].value == fes_name):
                visitor = np.array([i[1].value, i[2].value, i[3].value, i[4].value, i[5].value, i[6].value, i[7].value, i[8].value, i[9].value, i[10].value, i[11].value])
                print(visitor)
                break
    fig = plt.gcf()
    plt.plot(year, visitor,marker = 'o')
    plt.title('yearly visitor')
    plt.draw()
    file_name = "C:\\Users\\samsung\\Desktop\\2020-1\\데이터캡스톤디자인\\지역축제 방문객 예측\\프로젝트\\git\\web\\fes_visitor\\static\\fes_visitor.png"
    fig.savefig(file_name, dpi=fig.dpi)

def crawling_weather(location_code, year, month):
    #혹여나 이전에 에 있던 파일 지우기
    os.remove("C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/weather_django.txt")
    url = "https://www.weather.go.kr/weather/climate/past_cal.jsp?stn={}&yy={}&mm={}&obs=1&x=18&y=6".format(location_code, year, month)

    #크롬드라이버 설치 위치를 입력해준다. 
    driver = wd.Chrome("C:\/chromedriver")
    driver.get(url)
    time.sleep(2)  #페이지가 완전히 로드될 때까지 기다린다.
    
    #2)-1 시작월의 날씨 정보를 크롤링한다. 
    data = []
    table = driver.find_element_by_css_selector('table.table_develop')
    contents = table.find_element_by_tag_name('tbody')
    contents = contents.text
    
    # 날씨 정보를 weather.txt에 저장한다. 
    contents = contents.strip()
    weather_file = open("C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/weather_django.txt", 'w')
    weather_file.write(contents)
    weather_file.close()
    driver.close()

def start_weather(infos, start_day):
    f= open("C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/weather_django.txt", 'r')
    s= f.readlines()
    days = []
    lis = []
    index = 100
    info_count = 0 
    for_dic = []
    count = -1

    ## 시작일 이후부터 저장
    for i in s:
        if (start_day in i):
            i = i.strip('\n')
            days = i.split(" ")
            #위치 찾기
            index = days.index(start_day + "일") 
            print(index)
            count = -1

        if (index != 100):
            count += 1

        if(count > index*5):
            if("일 " not in i):
                for_dic = []
                data = i.strip('\n')
                print(data)
                data = data.strip()
                print(data)
                data = data.split(":")
                print(data)
                for_dic.append(data)
                print(for_dic)
                for_dic = dict(for_dic)
                lis.append(for_dic)
                print(lis)
                info_count += 1
                if(info_count == 5):
                    infos.append(lis)
                    info_count = 0
                    lis = []
        f.close()
    return infos

def end_weather(infos, end_day):
    f= open("C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/weather_django.txt", 'r')
    s= f.readlines()
    days = [] #일들을 담는 리스트
    lis = [] #행사 하나를 의미하는 리스트
    for_dic = [] # 딕셔너리로 만들기 위한 리스트
    index = 100 
    count = -1
    info_count = 0 

    ## 만약에 종료일이 3일이다
    for i in s:
        if (end_day in i):
            i = i.rstrip('\n')
            days = i.split(" ")
            print(days)
            #위치 찾기
            index = days.index(end_day+"일") 
            print("index: ", index)
            break
    info_count = 0
    for again in s:
        if(len(infos) == 3):
            break
        else:
            if("일 " not in again):
                for_dic = []
                again = again.rstrip('\n')
                print(again)
                again = again.strip()
                print(again)
                again = again.split(':')
                for_dic.append(again)
                lis.append(dict(for_dic))
                print(lis)
                info_count += 1
                print(info_count)
                if(info_count == 5):
                    infos.append(lis)
                    info_count = 0
                    lis = []
    f.close()
    return infos


def same_weather(infos, period):
    f= open("C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/weather_django.txt", 'r')
    s= f.readlines()
    days = [] #일들을 담는 리스트
    lis = [] #행사 하나를 의미하는 리스트
    for_dic = [] # 딕셔너리로 만들기 위한 리스트
    index = 100 
    count = -1
    info_count = 0 

    info_count = 0
    for again in s:
        if(len(infos) == period):
            break
        else:
            if("일 " not in again):
                for_dic = []
                again = again.rstrip('\n')
                print(again)
                again = again.strip()
                print(again)
                again = again.split(':')
                for_dic.append(again)
                lis.append(dict(for_dic))
                print(lis)
                info_count += 1
                print(info_count)
                if(info_count == 5):
                    infos.append(lis)
                    info_count = 0
                    lis = []
    f.close()
    return infos

def calc(infos):
    weather_info = []
    sum_avg_temp = 0.0 
    sum_high_temp = 0.0 
    sum_low_temp = 0.0
    sum_avg_cloud = 0.0  
    sum_rain = 0.0 
    weather = ""
    for day in infos:
        try:
            sum_avg_temp += float((day[0].get('평균기온')).replace("℃", ""))
            sum_high_temp += float((day[1].get('최고기온')).replace("℃", ""))
            sum_low_temp += float((day[2].get('최저기온')).replace("℃", ""))
            if((day[3].get('평균운량')) == ' -'):
                sum_avg_cloud += 0
            else:      
                sum_avg_cloud += float((day[3].get('평균운량')))

            if((day[4].get('일강수량')).replace("mm", "") == ' -'):
                sum_rain += 0
            else:
                sum_rain += float((day[4].get('일강수량')).replace("mm", ""))
        except:
            weather = "None\n"

    if(weather != "None\n"):       
        #축제별 avg_temp, high_temp, low_temp, avg_cloud, rain
        count = len(infos)
        avg_temp = round(sum_avg_temp/count, 2)
        high_temp = round(sum_high_temp/count)
        low_temp = round(sum_low_temp/count, 2)
        avg_cloud = round(sum_avg_cloud/count)
        rain = round(sum_rain/count, 2)   
        weather =  str(avg_temp) + ", " + str(high_temp) + ", " + str(low_temp) + ", " + str(avg_cloud) + ", " +str(rain)+ "\n"

    f = open("C:/Users/samsung/Desktop/2020-1/데이터캡스톤디자인/지역축제 방문객 예측/프로젝트/weather_info_django.txt", 'a')
    f.write(weather)
    f.close()
    return avg_temp, high_temp, low_temp, avg_cloud, rain

def avg_weather_crawling(year, location_code, start_month, 평균기온):
    avg_weather = []
    url = "https://www.weather.go.kr/weather/climate/past_table.jsp?stn={}&yy={}&obs=07&x=21&y=6".format(location_code, year)
    #크롬드라이버 설치 위치를 입력해준다. 
    driver = wd.Chrome("C:\chromedriver.exe")
    #driver = wd.Chrome("C:\Users\samsung\Downloads\chromedriver_win32\chromedriver.exe")
    driver.get(url)
    time.sleep(2)  #페이지가 완전히 로드될 때까지 기다린다.

    #2)-1 시작월의 날씨 정보를 크롤링한다. 
    data = []
    table = driver.find_element_by_css_selector('table.table_develop')
    contents = table.find_element_by_tag_name('tbody')
    contents = contents.text
    a = contents.find("평균")
    c = ((contents[a+3:len(contents)].split(" ")))
    for i in c:
        avg_weather.append(i)
    return abs(float(avg_weather[start_month-1]) - 평균기온)

def low_weather_crawling(year, location_code, start_month,  최저기온):
    avg_weather = []
    url = "https://www.weather.go.kr/weather/climate/past_table.jsp?stn={}&yy={}&obs=10&x=23&y=10".format(location_code, year)
    #크롬드라이버 설치 위치를 입력해준다. 
    driver = wd.Chrome("C:\chromedriver.exe")
    #driver = wd.Chrome("C:\Users\samsung\Downloads\chromedriver_win32\chromedriver.exe")
    driver.get(url)
    time.sleep(2)  #페이지가 완전히 로드될 때까지 기다린다.

    #2)-1 시작월의 날씨 정보를 크롤링한다. 
    data = []
    table = driver.find_element_by_css_selector('table.table_develop')
    contents = table.find_element_by_tag_name('tbody')
    contents = contents.text
    a = contents.find("평균")
    c = ((contents[a+3:len(contents)].split(" ")))
    for i in c:
        avg_weather.append(i)
    driver.close()
    return abs(float(avg_weather[start_month-1]) - 최저기온)

def high_weather_crawling(year, location_code, start_month, 최고기온):
    avg_weather = []
    url = "https://www.weather.go.kr/weather/climate/past_table.jsp?stn={}&yy={}&obs=08&x=13&y=15".format(location_code, year)
    #크롬드라이버 설치 위치를 입력해준다. 
    driver = wd.Chrome("C:\chromedriver.exe")
    #driver = wd.Chrome("C:\Users\samsung\Downloads\chromedriver_win32\chromedriver.exe")
    driver.get(url)
    time.sleep(2)  #페이지가 완전히 로드될 때까지 기다린다.

    #2)-1 시작월의 날씨 정보를 크롤링한다. 
    data = []
    table = driver.find_element_by_css_selector('table.table_develop')
    contents = table.find_element_by_tag_name('tbody')
    contents = contents.text
    a = contents.find("평균")
    c = ((contents[a+3:len(contents)].split(" ")))
    for i in c:
        avg_weather.append(i)
    driver.close()
    return abs(float(avg_weather[start_month-1]) - 최고기온)


def select_page(request):
    print(request.method)
    if request.method == "POST":
        form = Form(request.POST)
        # form의 내용이 유효하다면 DB에 저장한다. 
        options = Option.objects.all()
        results= Results.objects.all()    
        for option in options:
            option.delete()
        for result in results:
            result.delete()

        fes_name = request.POST.get('fes_name')
        print(fes_name)
        start_date = request.POST.get('start_date')
        print(start_date)
        end_date = request.POST.get('end_date')

        
        print(end_date)

        start_year = int(start_date[0:4])
        start_month = int(start_date[5:7])
        start_day = int(start_date[8:11])

        end_year = int(end_date[0:4])
        end_month = int(end_date[5:7])
        end_day = int(end_date[8:11])

        option_list = Option(fes_name = fes_name, fes_start_year = start_year, fes_start_month = start_month, fes_start_day= start_day, fes_end_year= end_year, fes_end_month = end_month, fes_end_day = end_day)
        option_list.save()

        day_term = calc_day(start_year, start_month, start_day, end_year, end_month, end_day)
        print(day_term)
        KTX, location_code, neutral, positive = calc_else(fes_name)

        infos = []
        ## 날씨 정보 수집
        if start_month != end_month:
            # 2)-1 시작월을 크롤링한다. (start_month)
            crawling_weather(location_code, start_year, start_month)
            # 2)-2시작일부터 해당 월의 마지막까지 가져온다.
            start_weather(infos, start_day)
            # 2)-3 종료 월을 크롤링한다.
            crawling_weather(location_code, start_year, end_month)
            # 2)-4 종료 일 전까의 데이터를 가져온다.
            end_weather(infos, end_day)
            calc(infos)

        else:
            # 3)-1 행사 월을 크롤링한다. (start_month)
            crawling_weather(location_code, start_year, start_month)
            # 3)-3 시작일부터 행사 일 수까지 데이터를 크롤링한다. 
            same_weather(infos, day_term)
            평균기온, 최고기온, 최저기온, 평균운량, 평균강수량 = calc(infos)

        평균기온절대값 = avg_weather_crawling(2019, location_code, start_month, 평균기온)
        최저기온절대값 = low_weather_crawling(2019, location_code, start_month, 최저기온)
        최고기온절대값 = high_weather_crawling(2019, location_code, start_month, 최고기온)
        
        test_value = df( data = {'기간': [day_term], "KTX역": [KTX],  '평균기온 절대값': [평균기온절대값], '최저기온 절대값': [최저기온절대값], '최고기온 절대값': [최고기온절대값],'평균운량': [평균운량], 'positive': [positive], 'neutral': [neutral]})
        

        regression_result = regression_model(test_value)
        bass_result = bass_model(fes_name, start_year)
        final_result = hybrid_model(regression_result, bass_result)
        result_list = Results(total_visitor = final_result, 평균기온 = 평균기온, 최저기온 = 최저기온, 최고기온= 최고기온)
        result_list.save()
     
        draw_visitor(fes_name)

        options = Option.objects.all()
        results= Results.objects.all()
        return render(request, 'result_page.html' , {'options': options, 'results': results})
    else:
        print("something wrong")
        form = Form()
        return render(request, 'select_page.html', {'form':form})
    
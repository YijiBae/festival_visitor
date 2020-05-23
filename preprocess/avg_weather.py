##평균 기온 크롤링
import requests
from bs4 import BeautifulSoup
from selenium import webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
from openpyxl import load_workbook
from collections import Counter
import os
from openpyxl import load_workbook
지역코드 =[]
지역이름 = []
Jen = []
Feb = []
Mar = []
Apr = []
May = []
Jun = []
Jul = []
Aug = []
Sep = []
Oct = []
Nov = []
Dec = []

#평균기온 url: https://www.weather.go.kr/weather/climate/past_table.jsp?stn={}&yy={}&obs=07&x=30&y=2
#최저기온 url: https://www.weather.go.kr/weather/climate/past_table.jsp?stn={}&yy={}&obs=10&x=19&y=15
#최고기온 url: https://www.weather.go.kr/weather/climate/past_table.jsp?stn={}&yy={}&obs=08&x=4&y=10

f = open("지역코드.txt", 'r', encoding = 'utf-8')
name = []
code = []
lines = f.readlines()
for line in lines:
    date = line.split(",")
    name.append(date[0])
    code.append(date[1].strip())
    
for i in range(0, len(code)):
    location_code = code[i]
    print(location_code)
    year = 2017
    url = "https://www.weather.go.kr/weather/climate/past_table.jsp?stn={}&yy={}&obs=08&x=4&y=10".format(location_code, year)
    #url  = "https://www.weather.go.kr/weather/climate/past_table.jsp?stn=108&yy=2009&obs=07&x=20&y=3"
    #크롬드라이버 설치 위치를 입력해준다. 
    driver = wd.Chrome("C:\chromedriver.exe")
    #driver = wd.Chrome("C:\Users\samsung\Downloads\chromedriver_win32\chromedriver.exe")
    driver.get(url)
    time.sleep(2)  #페이지가 완전히 로드될 때까지 기다린다.

    #2)-1 시작월의 날씨 정보를 크롤링한다. 
    data = []
    table = driver.find_element_by_css_selector('table.table_develop')
    contents = table.find_element_by_tag_name('tbody')
    contents = contents.text
    a = contents.find("평균")
    c = ((contents[a+3:len(contents)].split(" ")))
    지역이름.append(name[i])
    지역코드.append(location_code)
    for i in range(len(c)):
        if(i+1 == 1):
            Jen.append(c[i])
        elif(i+1 == 2):
            Feb.append(c[i])
        elif(i+1 == 3):
            Mar.append(c[i])
        elif(i+1 == 4):
            Apr.append(c[i])
        elif(i+1 == 5):
            May.append(c[i])
        elif(i+1 == 6):
            Jun.append(c[i])
        elif(i+1 == 7):
            Jul.append(c[i])
        elif(i+1 == 8):
            Aug.append(c[i])
        elif(i+1 == 9):
            Sep.append(c[i])
        elif(i+1 == 10):
            Oct.append(c[i])
        elif(i+1 == 11):
            Nov.append(c[i])
        elif(i+1 == 12):
            Dec.append(c[i])
    driver.close()
for i in 지역이름:
    print(i)
print("===============지역코드=======================")
for i in 지역코드:
    print(i)
print("==================1====================")
for i in Jen:
    print(i)
print("===================2===================")
for i in Feb:
    print(i)
print("====================3==================")
for i in Mar:
    print(i)
print("=====================4=================")
for i in Apr:
    print(i)
print("======================5================")
for i in May:
    print(i)
print("=======================6===============")
for i in Jun:
    print(i)
print("========================7==============")
for i in Jul:
    print(i)
print("=========================8=============")
for i in Aug:
    print(i)
print("==========================9============")
for i in Sep:
    print(i)
print("===========================10===========")
for i in Oct:
    print(i)
print("=============================11=========")
for i in Nov:
    print(i)
print("===============================12=======")
for i in Dec:
    print(i)


# 평균값-실제값 
def get_abs_weather():
    #1) 엑셀을 열어 행사 데이터를 꺼낸다. 
    lists = load_workbook('data_fesitival.xlsx')
    avg_wea = load_workbook('평균날씨.xlsx')
    year = ["2017년"]

    #['2007년', '2008년', '2009년', '2010년', '2011년', '2012년', '2013년', '2014년', '2015년', '2016년', '2017년']

    #행사 데이터 중 행사일을 읽는다. 
    for y in year:
        sheet = lists[y]
        weather_sheet = avg_wea[y]
        count_for_date = 1
        for i in sheet:
            count_for_date += 1
            # 셋째행까지는 무의미한 데이터가 삽입되어 있어 생략하기 위함이다. 
            if(count_for_date > 1 and i[8].value != None):
                location = i[2]                    # 서울  
                
                date = i[8].value                  #"2.27~5.6"
                date = date.split(".")            #["2", "27~5", "6"]
                start_month = date[0]
                
                for j in sheet:
                    for name in range(82):
                        if(j[name] in location or location in j[name]):
                            print(weather_sheet.cell(row=name,column=start_month+2).value)
                            break